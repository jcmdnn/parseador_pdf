import openpyxl as xl
import os
import pdfplumber

# busca arquivos PDF
for arquivo in os.listdir('pdfs'):

    if arquivo.lower().endswith('.pdf'):
        try:
            # abrindo arquivo excel
            excel = xl.load_workbook('Base de Dados Inspeções.xlsx')
            aba_selecionada = excel.active
            linha_inicial = len(aba_selecionada['A']) + 1
            # ler arquivo pdf e extrair os dados
            pdf = pdfplumber.open(f'pdfs\\{arquivo}')
            pagina = pdf.pages[0]
            dados = pagina.extract_table()

            # iterando sobre os dados do arquivo pdf, inserindo no excel
            for indice, dado in enumerate(dados[1:], start=linha_inicial):

                if dado[0] == '':
                    pass
                else:
                    aba_selecionada.cell(row=indice, column=1).value = dado[0]
                    aba_selecionada.cell(row=indice, column=2).value = dado[1]
                    aba_selecionada.cell(row=indice, column=3).value = dado[2]
                    aba_selecionada.cell(row=indice, column=4).value = dado[3]
                    aba_selecionada.cell(row=indice, column=5).value = dado[4]
            pdf.close()
            excel.save('Base de Dados Inspeções.xlsx')
            excel.close()
        except Exception as e:
            with open('log_erros.txt', 'a') as log:
                log.write(f'Erro ao extrair informações do arquivo {arquivo}.\n Erro: {e}')
    else:
        with open('log_erros.txt', 'a') as log:
            log.write(f'O arquivo {arquivo} não é um PDF válido.\n')
